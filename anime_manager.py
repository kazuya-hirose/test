import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime
import os
import glob
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.chrome.service import Service
from bs4 import BeautifulSoup
import time
import pykakasi
import re


def find_latest_anime_file():
    """最新のアニメデータExcelファイルを検出"""
    files = glob.glob("anime_data_*.xlsx")
    if files:
        # 最新ファイルを取得
        latest_file = max(files, key=os.path.getctime)
        return latest_file
    return None


def extract_hiragana_first_char(text):
    """テキストの最初の文字をひらがな1文字に統一"""
    if not text:
        return ""
    
    first_char = text[0]
    
    # ひらがなマッピング辞書
    katakana_to_hiragana = {
        'ア': 'あ', 'イ': 'い', 'ウ': 'う', 'エ': 'え', 'オ': 'お',
        'カ': 'か', 'キ': 'き', 'ク': 'く', 'ケ': 'け', 'コ': 'こ',
        'ガ': 'が', 'ギ': 'ぎ', 'グ': 'ぐ', 'ゲ': 'げ', 'ゴ': 'ご',
        'サ': 'さ', 'シ': 'し', 'ス': 'す', 'セ': 'せ', 'ソ': 'そ',
        'ザ': 'ざ', 'ジ': 'じ', 'ズ': 'ず', 'ゼ': 'ぜ', 'ゾ': 'ぞ',
        'タ': 'た', 'チ': 'ち', 'ツ': 'つ', 'テ': 'て', 'ト': 'と',
        'ダ': 'だ', 'ヂ': 'ぢ', 'ヅ': 'づ', 'デ': 'で', 'ド': 'ど',
        'ナ': 'な', 'ニ': 'に', 'ヌ': 'ぬ', 'ネ': 'ね', 'ノ': 'の',
        'ハ': 'は', 'ヒ': 'ひ', 'フ': 'ふ', 'ヘ': 'へ', 'ホ': 'ほ',
        'バ': 'ば', 'ビ': 'び', 'ブ': 'ぶ', 'ベ': 'べ', 'ボ': 'ぼ',
        'パ': 'ぱ', 'ピ': 'ぴ', 'プ': 'ぷ', 'ペ': 'ぺ', 'ポ': 'ぽ',
        'マ': 'ま', 'ミ': 'み', 'ム': 'む', 'メ': 'め', 'モ': 'も',
        'ヤ': 'や', 'ユ': 'ゆ', 'ヨ': 'よ',
        'ラ': 'ら', 'リ': 'り', 'ル': 'る', 'レ': 'れ', 'ロ': 'ろ',
        'ワ': 'わ', 'ヲ': 'を', 'ン': 'ん',
        'ッ': 'っ', 'ャ': 'ゃ', 'ュ': 'ゅ', 'ョ': 'ょ', 'ァ': 'ぁ',
        'ィ': 'ぃ', 'ゥ': 'ぅ', 'ェ': 'ぇ', 'ォ': 'ぉ'
    }
    
    # 英字からひらがなへのマッピング（ローマ字表記の最初の文字）
    alphabet_to_hiragana = {
        'A': 'あ', 'a': 'あ',
        'B': 'び', 'b': 'び',
        'C': 'し', 'c': 'し',
        'D': 'で', 'd': 'で',
        'E': 'え', 'e': 'え',
        'F': 'ふ', 'f': 'ふ',
        'G': 'ぎ', 'g': 'ぎ',
        'H': 'は', 'h': 'は',
        'I': 'い', 'i': 'い',
        'J': 'じ', 'j': 'じ',
        'K': 'か', 'k': 'か',
        'L': 'ら', 'l': 'ら',
        'M': 'ま', 'm': 'ま',
        'N': 'ん', 'n': 'ん',
        'O': 'お', 'o': 'お',
        'P': 'ぱ', 'p': 'ぱ',
        'Q': 'く', 'q': 'く',
        'R': 'ら', 'r': 'ら',
        'S': 'す', 's': 'す',
        'T': 'た', 't': 'た',
        'U': 'う', 'u': 'う',
        'V': 'ぶ', 'v': 'ぶ',
        'W': 'わ', 'w': 'わ',
        'X': 'え', 'x': 'え',
        'Y': 'や', 'y': 'や',
        'Z': 'ぜ', 'z': 'ぜ'
    }
    
    # ひらがなの場合はそのまま返す
    if 'あ' <= first_char <= 'ん':
        return first_char
    
    # カタカナの場合は変換
    if first_char in katakana_to_hiragana:
        return katakana_to_hiragana[first_char]
    
    # 英字の場合は変換
    if first_char in alphabet_to_hiragana:
        return alphabet_to_hiragana[first_char]
    
    # 記号や特殊文字の場合
    if first_char in '「『【':
        # 次の文字を見る
        if len(text) > 1:
            return extract_hiragana_first_char(text[1:])
    
    # 数字の場合
    if first_char.isdigit():
        return 'す'  # 数字は「すう」のす
    
    # 漢字の場合はpykakasiで変換
    try:
        kakasi = pykakasi.kakasi()
        
        # 複合漢字パターンを抽出（連続した複数の漢字）
        # 最初の漢字から始まる連続した漢字をまとめる
        kanji_count = 0
        for i, char in enumerate(text):
            if '\u4e00' <= char <= '\u9fff':  # 漢字の範囲
                kanji_count += 1
            else:
                break
        
        # 複合漢字がある場合は3文字までを試す
        if kanji_count >= 2:
            # 複合漢字を複数まとめて処理（最大3文字）
            for length in [min(3, kanji_count), 2]:
                compound = text[:length]
                result = kakasi.convert(compound)
                if result and len(result) > 0:
                    # 複合要素として認識されたかチェック
                    first_elem = result[0]
                    if len(first_elem.get('orig', '')) >= 2:
                        # 複合要素として認識された場合
                        hira = first_elem.get('hira', '')
                        if hira:
                            return hira[0]
            
            # 複合漢字が分割された場合は、複数要素を統合してみる
            result = kakasi.convert(text[:kanji_count])
            if result and len(result) > 0:
                # すべての要素のひらがなを連結して最初の文字を取得
                all_hira = ''
                for elem in result:
                    hira = elem.get('hira', '')
                    if hira and '\u4e00' <= elem.get('orig', '')[0] <= '\u9fff':
                        all_hira += hira
                if all_hira:
                    return all_hira[0]
        
        # 単一の漢字の場合
        result = kakasi.convert(first_char)
        if result and len(result) > 0:
            hira = result[0].get('hira', '')
            if hira:
                return hira[0]  # 最初の文字を返す
    except:
        pass
    
    # その他の場合は「そ」
    return 'そ'


def parse_theme_song(theme_text):
    """主題歌テキストからOP/ED情報を抽出"""
    op_title = ''
    op_artist = ''
    ed_title = ''
    ed_artist = ''
    
    # OP情報を抽出: OP：「{title}」{artist}
    if 'OP：' in theme_text:
        op_part = theme_text.split('OP：')[1]
        if '「' in op_part and '」' in op_part:
            op_title = op_part.split('「')[1].split('」')[0]
            # 「」の後のテキストがアーティスト
            after_title = op_part.split('」')[1]
            if 'ED：' in after_title:
                op_artist = after_title.split('ED：')[0].strip()
            else:
                op_artist = after_title.strip()
    
    # ED情報を抽出: ED：「{title}」{artist}
    if 'ED：' in theme_text:
        ed_part = theme_text.split('ED：')[1]
        if '「' in ed_part and '」' in ed_part:
            ed_title = ed_part.split('「')[1].split('」')[0]
            # 「」の後のテキストがアーティスト
            after_title = ed_part.split('」')[1]
            ed_artist = after_title.strip()
    
    return op_title, op_artist, ed_title, ed_artist


def extract_anime_details(soup, anime_title):
    """テーブルからアニメの詳細情報を抽出"""
    details = {
        '放送年・期': '',
        'OP曲': '',
        'OP歌手': '',
        'ED曲': '',
        'ED歌手': ''
    }
    
    # すべてのテーブルを確認
    tables = soup.find_all('table')
    for table in tables:
        # テーブルのテキストに作品名が含まれているか確認
        table_text = table.get_text(strip=True)
        if anime_title not in table_text:
            continue
        
        # このテーブルから情報を抽出
        rows = table.find_all('tr')
        for row in rows:
            cells = row.find_all(['td', 'th'])
            if len(cells) >= 2:
                label = cells[0].get_text(strip=True)
                value = cells[1].get_text(strip=True)
                
                if '主題歌' in label:
                    op_title, op_artist, ed_title, ed_artist = parse_theme_song(value)
                    details['OP曲'] = op_title
                    details['OP歌手'] = op_artist
                    details['ED曲'] = ed_title
                    details['ED歌手'] = ed_artist
                
                elif any(key in label for key in ['公開開始年', '放送開始', '季節']):
                    details['放送年・期'] = value
        
        # 見つかったら処理終了
        if details['OP曲'] or details['放送年・期']:
            break
    
    return details


def scrape_anime_from_url(url):
    """SeleniumでURLからアニメ情報をスクレイピング（詳細情報付き）"""
    print(f"\nURL: {url}")
    print("ページを読み込み中...")
    
    options = webdriver.ChromeOptions()
    options.add_argument("--no-sandbox")
    options.add_argument("--disable-dev-shm-usage")
    options.add_argument("--disable-blink-features=AutomationControlled")
    options.add_argument("user-agent=Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36")
    
    driver = None
    try:
        service = Service(ChromeDriverManager().install())
        driver = webdriver.Chrome(service=service, options=options)
        driver.get(url)
        
        # ページの読み込みを待つ
        time.sleep(3)
        
        # JavaScriptが実行されるのを待つ
        WebDriverWait(driver, 10).until(
            lambda d: len(d.find_elements(By.TAG_NAME, "a")) > 10
        )
        
        html = driver.page_source
        soup = BeautifulSoup(html, 'html.parser')
        
        anime_list = []
        
        # 目次のリストから作品名を抽出
        lists = soup.find_all(['ul', 'ol'])
        
        for lst in lists:
            links = lst.find_all('a', href=lambda x: x and '#' in x)
            
            for link in links:
                title_text = link.get_text(strip=True)
                
                # タイトルが空ではなく、一定の長さがある場合
                if title_text and len(title_text) > 1:
                    # 重複チェック
                    if not any(a['作品名'] == title_text for a in anime_list):
                        # 詳細情報を抽出
                        details = extract_anime_details(soup, title_text)
                        
                        anime_list.append({
                            '作品名': title_text,
                            '頭文字': extract_hiragana_first_char(title_text),
                            '放送年・期': details['放送年・期'],
                            'OP曲': details['OP曲'],
                            'OP歌手': details['OP歌手'],
                            'ED曲': details['ED曲'],
                            'ED歌手': details['ED歌手']
                        })
        
        # 見つからない場合は別の方法を試す
        if not anime_list:
            list_divs = soup.find_all(['div', 'nav'], class_=lambda x: x and 'list' in x.lower())
            for div in list_divs:
                links = div.find_all('a', href=lambda x: x and '#' in x)
                for link in links:
                    title_text = link.get_text(strip=True)
                    if title_text and len(title_text) > 1:
                        if not any(a['作品名'] == title_text for a in anime_list):
                            details = extract_anime_details(soup, title_text)
                            anime_list.append({
                                '作品名': title_text,
                                '頭文字': extract_hiragana_first_char(title_text),
                                '放送年・期': details['放送年・期'],
                                'OP曲': details['OP曲'],
                                'OP歌手': details['OP歌手'],
                                'ED曲': details['ED曲'],
                                'ED歌手': details['ED歌手']
                            })
        
        print(f"✓ {len(anime_list)}件のアニメ情報を抽出しました")
        return anime_list
    
    except Exception as e:
        print(f"✗ スクレイピング中にエラーが発生しました: {e}")
        return []
    
    finally:
        if driver:
            driver.quit()


def create_anime_template(filename=None):
    """アニメ情報入力用のExcel雛形を作成"""
    if not filename:
        filename = f"anime_data_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    # ワークブック作成
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "アニメ情報"
    
    # ヘッダー行
    headers = ['作品名', '頭文字', '放送年・期', 'OP曲', 'OP歌手', 'ED曲', 'ED歌手']
    ws.append(headers)
    
    # ヘッダースタイル
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # 列幅を調整
    ws.column_dimensions['A'].width = 25  # 作品名
    ws.column_dimensions['B'].width = 8   # 頭文字
    ws.column_dimensions['C'].width = 15  # 放送年・期
    ws.column_dimensions['D'].width = 25  # OP曲
    ws.column_dimensions['E'].width = 20  # OP歌手
    ws.column_dimensions['F'].width = 25  # ED曲
    ws.column_dimensions['G'].width = 20  # ED歌手
    
    # 行高さを調整
    ws.row_dimensions[1].height = 25
    
    # データ入力用の空行を30行作成
    for row in range(2, 32):
        for col in range(1, 8):
            cell = ws.cell(row=row, column=col)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            ws.row_dimensions[row].height = 20
    
    # 保存
    wb.save(filename)
    print(f"✓ Excel雛形を作成しました: {filename}")
    return filename


def add_anime_data(filename, anime_list):
    """新規Excelファイルにアニメデータを追加（既に雛形が作成されていることを想定）"""
    if not filename:
        return None
    
    try:
        wb = openpyxl.load_workbook(filename)
        ws = wb.active
        
        # 既に入力されている行を検出
        start_row = 2
        for row in range(2, ws.max_row + 1):
            if not ws[f'A{row}'].value:
                start_row = row
                break
        
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # データを追加
        for idx, anime in enumerate(anime_list):
            row = start_row + idx
            ws[f'A{row}'] = anime.get('作品名', '')
            ws[f'B{row}'] = anime.get('頭文字', '')
            ws[f'C{row}'] = anime.get('放送年・期', '')
            ws[f'D{row}'] = anime.get('OP曲', '')
            ws[f'E{row}'] = anime.get('OP歌手', '')
            ws[f'F{row}'] = anime.get('ED曲', '')
            ws[f'G{row}'] = anime.get('ED歌手', '')
            
            # セルのボーダーを設定
            for col in range(1, 8):
                cell = ws.cell(row=row, column=col)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        
        wb.save(filename)
        print(f"✓ {len(anime_list)}件のデータを追加しました")
        return filename
    
    except Exception as e:
        print(f"エラー: {e}")
        return None


def interactive_input():
    """インタラクティブにアニメ情報を入力"""
    anime_list = []
    
    print("\n" + "=" * 60)
    print("アニメ情報入力ツール")
    print("=" * 60)
    
    while True:
        print("\n" + "-" * 60)
        print(f"作品 #{len(anime_list) + 1}")
        print("-" * 60)
        
        title = input("作品名（終了: 'done'）: ").strip()
        if title.lower() == 'done':
            break
        
        if not title:
            print("作品名は必須です")
            continue
        
        first_char = input("頭文字（ひらがな）: ").strip()
        broadcast = input("放送年・期（例: 2023年冬）: ").strip()
        op_title = input("OP曲: ").strip()
        op_artist = input("OP歌手: ").strip()
        ed_title = input("ED曲: ").strip()
        ed_artist = input("ED歌手: ").strip()
        
        anime_list.append({
            '作品名': title,
            '頭文字': first_char,
            '放送年・期': broadcast,
            'OP曲': op_title,
            'OP歌手': op_artist,
            'ED曲': ed_title,
            'ED歌手': ed_artist
        })
        
        print(f"✓ 追加しました（現在 {len(anime_list)}件）")
    
    return anime_list


def main():
    print("\n" + "=" * 60)
    print("アニメ情報管理ツール（Excel出力 + URL スクレイピング）")
    print("=" * 60)
    
    # メニュー表示
    while True:
        print("\n【メニュー】")
        print("1. URLからスクレイピング → 新規Excelファイルに保存")
        print("2. 手動入力 → 新規Excelファイルに保存")
        print("3. 終了")
        
        choice = input("\n選択してください（1-3）: ").strip()
        
        if choice == '1':
            # URLからスクレイピング
            url = input("スクレイピング対象のURLを入力してください: ").strip()
            if not url.startswith(('http://', 'https://')):
                print("エラー: URLはhttp://またはhttps://で始まる必要があります")
                continue
            
            anime_list = scrape_anime_from_url(url)
            
            if anime_list:
                print("\n【スクレイピング結果】")
                for idx, anime in enumerate(anime_list, 1):
                    print(f"{idx}. {anime['作品名']} ({anime['頭文字']})")
                    if anime['OP曲']:
                        print(f"    OP: 「{anime['OP曲']}」{anime['OP歌手']}")
                    if anime['ED曲']:
                        print(f"    ED: 「{anime['ED曲']}」{anime['ED歌手']}")
                
                # ファイル名を自動生成
                # 放送年・期情報を取得
                broadcast_info = ""
                for anime in anime_list:
                    if anime['放送年・期']:
                        broadcast_info = anime['放送年・期'].replace('年', '').replace('アニメ', '')
                        break
                
                if broadcast_info:
                    filename = f"anime_data_{broadcast_info}_{len(anime_list)}件_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
                else:
                    filename = f"anime_data_{len(anime_list)}件_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
                
                # ファイル作成
                created_file = create_anime_template(filename)
                if created_file:
                    add_anime_data(created_file, anime_list)
                    print(f"✓ ファイルを作成しました: {created_file}")
        
        elif choice == '2':
            # 手動入力 → 新規ファイル作成
            anime_list = interactive_input()
            
            if anime_list:
                # ファイル名を自動生成
                broadcast_info = ""
                for anime in anime_list:
                    if anime['放送年・期']:
                        broadcast_info = anime['放送年・期'].replace('年', '').replace('アニメ', '')
                        break
                
                if broadcast_info:
                    filename = f"anime_data_{broadcast_info}_{len(anime_list)}件_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
                else:
                    filename = f"anime_data_{len(anime_list)}件_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
                
                # ファイル作成
                created_file = create_anime_template(filename)
                if created_file:
                    add_anime_data(created_file, anime_list)
                    print(f"✓ ファイルを作成しました: {created_file}")
        
        elif choice == '3':
            print("終了します。")
            break
        
        else:
            print("無効な選択です")


if __name__ == "__main__":
    main()
